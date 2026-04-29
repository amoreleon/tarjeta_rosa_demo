import os
import re
import json
import time
import io
import csv
import threading
from datetime import datetime, timezone

from flask import Flask, request, render_template, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from twilio.rest import Client

from flask_sqlalchemy import SQLAlchemy

load_dotenv()

# =========================
# Environment / Config
# =========================
ACCOUNT_SID = os.getenv("TWILIO_ACCOUNT_SID", "").strip()
AUTH_TOKEN = os.getenv("TWILIO_AUTH_TOKEN", "").strip()
DEFAULT_FLOW_SID = os.getenv("TWILIO_STUDIO_FLOW_SID", "").strip()
DEFAULT_FROM_NUMBER = os.getenv("TWILIO_FROM_NUMBER", "").strip()

DEMO_PASSCODE = (os.getenv("DEMO_PASSCODE", "") or "").strip()
FRONTEND_ORIGIN = (os.getenv("FRONTEND_ORIGIN", "") or "").strip()

STUDIO_WEBHOOK_SECRET = (os.getenv("STUDIO_WEBHOOK_SECRET", "") or "").strip()

MAX_CALLS_PER_REQUEST = int(os.getenv("MAX_CALLS_PER_REQUEST", "1000"))
CALLS_PER_SECOND = float(os.getenv("CALLS_PER_SECOND", "1"))

# Flex / TaskRouter
TWILIO_WORKSPACE_SID = os.getenv("TWILIO_WORKSPACE_SID", "").strip()
TWILIO_WORKFLOW_SID = os.getenv("TWILIO_WORKFLOW_SID", "WW002ca780443ec39460438835c4dc7bd1").strip()
TWILIO_TASK_CHANNEL = os.getenv("TWILIO_TASK_CHANNEL", "voice").strip()

DATABASE_URL = (os.getenv("DATABASE_URL", "") or "").strip()

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

if not DATABASE_URL:
    DATABASE_URL = "sqlite:///local.db"

# =========================
# App init
# =========================
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET", "change-me")

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

cors_origins = [FRONTEND_ORIGIN] if FRONTEND_ORIGIN else "*"
CORS(
    app,
    resources={
        r"/call": {"origins": cors_origins},
        r"/auth-check": {"origins": cors_origins},
        r"/api/*": {"origins": cors_origins},
        r"/flex-tasks": {"origins": cors_origins},
        r"/studio-webhook": {"origins": "*"},
    },
    methods=["GET", "POST", "OPTIONS"],
)

twilio_client = Client(ACCOUNT_SID, AUTH_TOKEN) if (ACCOUNT_SID and AUTH_TOKEN) else None


# =========================
# DB Models
# =========================
def utcnow():
    return datetime.now(timezone.utc)


class Launch(db.Model):
    __tablename__ = "launches"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime(timezone=True), default=utcnow, index=True, nullable=False)
    campaign_id = db.Column(db.String(128), index=True, nullable=False)
    flow_sid = db.Column(db.String(64), index=True, nullable=True)
    execution_sid = db.Column(db.String(64), index=True, nullable=True)
    to_number = db.Column(db.String(32), index=True, nullable=False)
    from_number = db.Column(db.String(32), nullable=True)


class Result(db.Model):
    __tablename__ = "results"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime(timezone=True), default=utcnow, index=True, nullable=False)
    campaign_id = db.Column(db.String(128), index=True, nullable=False)
    flow_sid = db.Column(db.String(64), index=True, nullable=True)
    execution_sid = db.Column(db.String(64), index=True, nullable=True)
    call_sid = db.Column(db.String(64), index=True, nullable=True)
    to_number = db.Column(db.String(32), index=True, nullable=True)
    answers_json = db.Column(db.Text, nullable=False, default="{}")
    raw_json = db.Column(db.Text, nullable=False, default="{}")


class CallEvent(db.Model):
    __tablename__ = "call_events"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime(timezone=True), default=utcnow, index=True, nullable=False)
    campaign_id = db.Column(db.String(128), index=True, nullable=False)
    flow_sid = db.Column(db.String(64), index=True, nullable=True)
    execution_sid = db.Column(db.String(64), index=True, nullable=True)
    call_sid = db.Column(db.String(64), index=True, nullable=True)
    to_number = db.Column(db.String(32), index=True, nullable=True)
    outcome = db.Column(db.String(32), index=True, nullable=False)
    raw_json = db.Column(db.Text, nullable=False, default="{}")


class FlexTask(db.Model):
    __tablename__ = "flex_tasks"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime(timezone=True), default=utcnow, index=True, nullable=False)
    campaign_id = db.Column(db.String(128), index=True, nullable=False)
    to_number = db.Column(db.String(32), index=True, nullable=False)
    task_sid = db.Column(db.String(64), index=True, nullable=True)
    worker_name = db.Column(db.String(128), nullable=True)
    # pending, assigned, completed, canceled
    status = db.Column(db.String(32), index=True, nullable=False, default="pending")
    completed_at = db.Column(db.DateTime(timezone=True), nullable=True)


with app.app_context():
    db.create_all()


# =========================
# Helpers
# =========================
def require_passcode(payload: dict):
    if not DEMO_PASSCODE:
        return None
    provided = (payload.get("passcode") or "").strip()
    if not provided:
        provided = (request.headers.get("X-CC-Passcode", "") or "").strip()
    if provided != DEMO_PASSCODE:
        return {"ok": False, "error": "Passcode inválido"}, 401
    return None


def normalize_mx_number(value: str):
    digits = re.sub(r"[^0-9]", "", str(value or ""))
    if not digits:
        return None
    if len(digits) == 10:
        digits = "52" + digits
    if len(digits) == 12 and digits.startswith("52"):
        return "+" + digits
    return None


def parse_numbers(raw):
    if raw is None:
        return [], []
    if isinstance(raw, list):
        lines = [str(x) for x in raw]
    else:
        lines = str(raw).splitlines()
    valid, invalid, seen = [], [], set()
    for line in lines:
        line = (line or "").strip()
        if not line:
            continue
        norm = normalize_mx_number(line)
        if not norm:
            invalid.append(line)
            continue
        if norm not in seen:
            seen.add(norm)
            valid.append(norm)
    return valid, invalid


def safe_json(obj):
    try:
        return json.dumps(obj, ensure_ascii=False)
    except Exception:
        return "{}"


def read_answers(payload: dict):
    answers = payload.get("answers")
    if isinstance(answers, dict):
        return answers
    out = {}
    for k, v in payload.items():
        if isinstance(k, str) and re.fullmatch(r"q\d+", k.strip().lower()):
            out[k.strip().lower()] = str(v)
    return out


def normalize_outcome(v: str):
    v = (v or "").strip().lower()
    if v in ("no_answer", "noanswer", "no-answer", "no answer"):
        return "no-answer"
    if v in ("busy",):
        return "busy"
    if v in ("failed", "call_failed", "call-failed", "error"):
        return "failed"
    if v in ("completed", "complete", "survey_result"):
        return "completed"
    if v in ("voicemail", "machine"):
        return "no-answer"
    return v or "unknown"


def clasificar_estado(outcome, duration=None, answers=None, audio_duration=30):
    """
    Clasifica el estado de una llamada de forma legible para el usuario.
    outcome: string de Twilio (completed, no-answer, busy, failed, canceled, etc.)
    duration: duración en segundos (int o None)
    answers: dict de respuestas IVR o None
    audio_duration: duración estimada del audio en segundos (default 30)
    """
    outcome = (outcome or "").strip().lower()
    answers = answers or {}
    has_answers = bool(answers) and any(
        v and str(v).strip() and str(v).strip().lower() != "sin respuesta"
        for v in answers.values()
    )

    try:
        dur = int(duration) if duration not in (None, "", "None") else None
    except (ValueError, TypeError):
        dur = None

    if outcome in ("no-answer", "no_answer", "noanswer"):
        return "No contestó"
    if outcome == "busy":
        return "Ocupado"
    if outcome == "failed":
        return "Falló"
    if outcome == "canceled":
        return "Cancelada"
    if outcome == "completed":
        if has_answers:
            return "Completó encuesta"
        if dur is not None:
            if dur >= (audio_duration - 5):
                return "Posible buzón"
            if dur >= 5:
                return "Contestó sin participar"
            return "Colgó rápido"
        return "Contestó sin participar"
    # pending o desconocido
    return "Sin resultado"


# Estados que vale la pena relanzar
ESTADOS_RELANZAR = {
    "No contestó", "Ocupado", "Posible buzón",
    "Contestó sin participar", "Colgó rápido", "Sin resultado"
}


def find_existing_call_event(call_sid: str, execution_sid: str):
    call_sid = call_sid or None
    execution_sid = execution_sid or None
    if call_sid:
        existing = CallEvent.query.filter(CallEvent.call_sid == call_sid).first()
        if existing:
            return existing
    if execution_sid:
        existing = CallEvent.query.filter(CallEvent.execution_sid == execution_sid).first()
        if existing:
            return existing
    return None


# =========================
# Routes — Core
# =========================
@app.get("/")
def index():
    return render_template("index.html", flow_sid=DEFAULT_FLOW_SID, from_number=DEFAULT_FROM_NUMBER)


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/auth-check")
def auth_check():
    payload = request.get_json(silent=True) or {}
    fail = require_passcode(payload)
    if fail:
        return fail
    return {"ok": True}


# =========================
# Routes — IVR Outbound
# =========================
def _launch_calls_background(app_ctx, campaign_id, flow_sid, from_number, to_numbers, extra_params):
    sleep_between = (1.0 / CALLS_PER_SECOND) if (CALLS_PER_SECOND and CALLS_PER_SECOND > 0) else 1.0
    with app_ctx:
        for to in to_numbers:
            try:
                params = {"campaignId": campaign_id, **extra_params}
                execution = twilio_client.studio.v2.flows(flow_sid).executions.create(
                    to=to, from_=from_number, parameters=params
                )
                db.session.add(Launch(
                    campaign_id=campaign_id,
                    flow_sid=flow_sid,
                    execution_sid=execution.sid,
                    to_number=to,
                    from_number=from_number
                ))
                db.session.commit()
            except Exception as e:
                print(f"[background] error to={to}: {e}", flush=True)
            time.sleep(sleep_between)
        print(f"[background] campaign={campaign_id} finished total={len(to_numbers)}", flush=True)


@app.post("/call")
def call():
    if not twilio_client:
        return {"ok": False, "error": "Twilio no configurado (credenciales faltantes)."}, 500

    payload = request.get_json(silent=True) or request.form.to_dict(flat=True) or {}
    fail = require_passcode(payload)
    if fail:
        return fail

    flow_sid = (payload.get("flow_sid") or payload.get("flowSid") or DEFAULT_FLOW_SID or "").strip()
    from_number = (payload.get("from_number") or payload.get("fromNumber") or DEFAULT_FROM_NUMBER or "").strip()

    if not flow_sid:
        return {"ok": False, "error": "Falta flow_sid."}, 400
    if not from_number:
        return {"ok": False, "error": "Falta from_number."}, 400

    campaign_id = (payload.get("campaign_id") or payload.get("campaignId") or "contact-center-v1").strip()
    extra_params = payload.get("extra_params") or payload.get("extraParams") or {}
    if isinstance(extra_params, str):
        try:
            extra_params = json.loads(extra_params) if extra_params.strip() else {}
        except Exception as e:
            return {"ok": False, "error": f"extra_params debe ser JSON válido: {e}"}, 400
    if not isinstance(extra_params, dict):
        extra_params = {}

    to_numbers_raw = payload.get("to_numbers") or payload.get("toNumbers") or ""
    to_numbers, invalid = parse_numbers(to_numbers_raw)

    if not to_numbers:
        return {"ok": False, "error": "No hay números válidos.", "invalid": invalid}, 400

    if MAX_CALLS_PER_REQUEST > 0 and len(to_numbers) > MAX_CALLS_PER_REQUEST:
        to_numbers = to_numbers[:MAX_CALLS_PER_REQUEST]

    ctx = app.app_context()
    t = threading.Thread(
        target=_launch_calls_background,
        args=(ctx, campaign_id, flow_sid, from_number, to_numbers, extra_params),
        daemon=True
    )
    t.start()

    return {
        "ok": True,
        "campaign_id": campaign_id,
        "flow_sid": flow_sid,
        "queued_count": len(to_numbers),
        "invalid_count": len(invalid),
        "invalid_preview": invalid[:50],
        "message": f"{len(to_numbers)} llamadas en proceso. Se lanzarán a {CALLS_PER_SECOND}/seg en background.",
    }


@app.post("/studio-webhook")
def studio_webhook():
    if STUDIO_WEBHOOK_SECRET:
        provided = (request.headers.get("X-Webhook-Secret", "") or "").strip()
        if provided != STUDIO_WEBHOOK_SECRET:
            return {"ok": False, "error": "Unauthorized"}, 401

    payload = request.get_json(silent=True) or request.form.to_dict(flat=True) or {}
    event = (payload.get("event") or "").strip().lower()
    campaign_id = (payload.get("campaignId") or payload.get("campaign_id") or "unknown").strip()
    flow_sid = (payload.get("flowSid") or payload.get("flow_sid") or "").strip()
    execution_sid = (payload.get("executionSid") or payload.get("execution_sid") or "").strip()
    call_sid = (payload.get("callSid") or payload.get("CallSid") or "").strip() or None
    to_number = (payload.get("to") or payload.get("To") or payload.get("contact") or "").strip() or None
    raw_outcome = (payload.get("callOutcome") or payload.get("call_outcome") or "").strip()

    app.logger.info("studio_webhook event=%s campaign=%s call=%s", event, campaign_id, call_sid)

    if event == "call_result":
        call_outcome = normalize_outcome(raw_outcome)
        existing = find_existing_call_event(call_sid, execution_sid)
        if existing:
            existing.outcome = call_outcome
            existing.raw_json = safe_json(payload)
            existing.campaign_id = campaign_id
            existing.flow_sid = flow_sid or existing.flow_sid
            existing.execution_sid = execution_sid or existing.execution_sid
            existing.to_number = to_number or existing.to_number
            if call_sid and not existing.call_sid:
                existing.call_sid = call_sid
        else:
            db.session.add(CallEvent(
                campaign_id=campaign_id, flow_sid=flow_sid or None,
                execution_sid=execution_sid or None, call_sid=call_sid or None,
                to_number=to_number or None, outcome=call_outcome, raw_json=safe_json(payload),
            ))
        db.session.commit()
        return {"ok": True}

    answers = read_answers(payload)
    db.session.add(Result(
        campaign_id=campaign_id, flow_sid=flow_sid or None,
        execution_sid=execution_sid or None, call_sid=call_sid or None,
        to_number=to_number or None, answers_json=safe_json(answers), raw_json=safe_json(payload),
    ))

    survey_outcome = normalize_outcome(raw_outcome or "completed")
    existing = find_existing_call_event(call_sid, execution_sid)
    if existing:
        existing.outcome = survey_outcome
        existing.raw_json = safe_json(payload)
        existing.campaign_id = campaign_id
        existing.flow_sid = flow_sid or existing.flow_sid
        existing.execution_sid = execution_sid or existing.execution_sid
        existing.to_number = to_number or existing.to_number
        if call_sid and not existing.call_sid:
            existing.call_sid = call_sid
    else:
        db.session.add(CallEvent(
            campaign_id=campaign_id, flow_sid=flow_sid or None,
            execution_sid=execution_sid or None, call_sid=call_sid or None,
            to_number=to_number or None, outcome=survey_outcome, raw_json=safe_json(payload),
        ))
    db.session.commit()
    return {"ok": True}


# =========================
# Routes — Flex Task Queue
# =========================
def _create_flex_tasks_background(app_ctx, campaign_id, to_numbers):
    """Crea Tasks en TaskRouter para cola de agentes Flex."""
    with app_ctx:
        created = 0
        errors = 0
        for to in to_numbers:
            try:
                attributes = json.dumps({
                    "phone": to,
                    "campaign_id": campaign_id,
                    "type": "outbound",
                    "name": to,
                })
                task = twilio_client.taskrouter.v1.workspaces(TWILIO_WORKSPACE_SID).tasks.create(
                    attributes=attributes,
                    workflow_sid=TWILIO_WORKFLOW_SID,
                    task_channel=TWILIO_TASK_CHANNEL,
                    timeout=86400  # 24 horas máximo en cola
                )
                db.session.add(FlexTask(
                    campaign_id=campaign_id,
                    to_number=to,
                    task_sid=task.sid,
                    status="pending"
                ))
                db.session.commit()
                created += 1
            except Exception as e:
                print(f"[flex-tasks] error to={to}: {e}", flush=True)
                errors += 1
            time.sleep(0.1)  # 10 tasks/seg
        print(f"[flex-tasks] campaign={campaign_id} created={created} errors={errors}", flush=True)


@app.post("/flex-tasks")
def create_flex_tasks():
    """Crea tareas en Flex TaskRouter para marcación saliente con agente."""
    if not twilio_client:
        return {"ok": False, "error": "Twilio no configurado."}, 500
    if not TWILIO_WORKSPACE_SID:
        return {"ok": False, "error": "Falta TWILIO_WORKSPACE_SID en variables de entorno."}, 500

    payload = request.get_json(silent=True) or {}
    fail = require_passcode(payload)
    if fail:
        return fail

    campaign_id = (payload.get("campaign_id") or payload.get("campaignId") or "flex-campaign").strip()
    to_numbers_raw = payload.get("to_numbers") or payload.get("toNumbers") or ""
    to_numbers, invalid = parse_numbers(to_numbers_raw)

    if not to_numbers:
        return {"ok": False, "error": "No hay números válidos.", "invalid": invalid}, 400

    if MAX_CALLS_PER_REQUEST > 0 and len(to_numbers) > MAX_CALLS_PER_REQUEST:
        to_numbers = to_numbers[:MAX_CALLS_PER_REQUEST]

    ctx = app.app_context()
    t = threading.Thread(
        target=_create_flex_tasks_background,
        args=(ctx, campaign_id, to_numbers),
        daemon=True
    )
    t.start()

    return {
        "ok": True,
        "campaign_id": campaign_id,
        "queued_count": len(to_numbers),
        "invalid_count": len(invalid),
        "message": f"{len(to_numbers)} tareas creándose en cola de Flex.",
    }


@app.get("/api/flex-tasks")
def get_flex_tasks():
    """Lista tareas Flex con estatus."""
    fail = require_passcode(request.args.to_dict())
    if fail:
        return fail

    campaign_id = (request.args.get("campaign_id") or "").strip()
    status_filter = (request.args.get("status") or "").strip()
    limit = int(request.args.get("limit", 200))

    q = FlexTask.query
    if campaign_id:
        q = q.filter(FlexTask.campaign_id == campaign_id)
    if status_filter:
        q = q.filter(FlexTask.status == status_filter)
    tasks = q.order_by(FlexTask.created_at.desc()).limit(limit).all()

    result = []
    for t in tasks:
        result.append({
            "id": t.id,
            "campaign_id": t.campaign_id,
            "to_number": t.to_number,
            "task_sid": t.task_sid,
            "status": t.status,
            "worker_name": t.worker_name,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "completed_at": t.completed_at.isoformat() if t.completed_at else None,
        })

    total = len(result)
    pending = sum(1 for r in result if r["status"] == "pending")
    assigned = sum(1 for r in result if r["status"] == "assigned")
    completed = sum(1 for r in result if r["status"] in ("completed", "canceled"))

    return {
        "ok": True,
        "tasks": result,
        "summary": {
            "total": total,
            "pending": pending,
            "assigned": assigned,
            "completed": completed,
        }
    }


@app.post("/api/flex-tasks/webhook")
def flex_task_webhook():
    """TaskRouter webhook — actualiza estatus de tareas cuando Twilio notifica."""
    payload = request.form.to_dict(flat=True) or request.get_json(silent=True) or {}

    task_sid = payload.get("TaskSid", "")
    event_type = (payload.get("EventType", "") or "").lower()
    worker_name = payload.get("WorkerName", "")
    assignment_status = (payload.get("TaskAssignmentStatus", "") or "").lower()

    print(f"[flex-webhook] TaskSid={task_sid} Event={event_type} Status={assignment_status} Worker={worker_name}", flush=True)

    if not task_sid:
        return {"ok": True}

    task = FlexTask.query.filter_by(task_sid=task_sid).first()
    if not task:
        return {"ok": True}

    # Manejar tanto EventType como TaskAssignmentStatus
    is_assigned = (
        assignment_status == "assigned"
        or event_type in ("reservation.accepted", "task.assigned")
    )
    is_completed = (
        assignment_status in ("completed", "canceled", "wrapping")
        or event_type in ("task.completed", "task.canceled", "reservation.completed", "reservation.wrapup")
    )

    if is_assigned and task.status == "pending":
        task.status = "assigned"
        task.worker_name = worker_name or task.worker_name
    elif is_completed:
        task.status = "completed"
        task.completed_at = utcnow()
        task.worker_name = worker_name or task.worker_name

    db.session.commit()
    return {"ok": True}


# =========================
# Routes — API consultas
# =========================
@app.get("/api/results")
def api_results():
    fail = require_passcode(request.args.to_dict(flat=True))
    if fail:
        return fail

    campaign_id = (request.args.get("campaign_id") or request.args.get("campaignId") or "").strip()
    flow_sid = (request.args.get("flow_sid") or request.args.get("flowSid") or "").strip()
    limit = int(request.args.get("limit", "200"))

    q = Result.query.order_by(Result.created_at.desc())
    if campaign_id:
        q = q.filter(Result.campaign_id == campaign_id)
    if flow_sid:
        q = q.filter(Result.flow_sid == flow_sid)

    rows = q.limit(min(limit, 1000)).all()
    out = []
    for r in rows:
        try:
            answers = json.loads(r.answers_json or "{}")
        except Exception:
            answers = {}
        out.append({
            "id": r.id,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "campaign_id": r.campaign_id,
            "flow_sid": r.flow_sid,
            "execution_sid": r.execution_sid,
            "call_sid": r.call_sid,
            "to": r.to_number,
            "answers": answers,
        })
    return {"ok": True, "count": len(out), "results": out}


@app.get("/api/summary")
def api_summary():
    fail = require_passcode(request.args.to_dict(flat=True))
    if fail:
        return fail

    campaign_id = (request.args.get("campaign_id") or request.args.get("campaignId") or "").strip()
    flow_sid = (request.args.get("flow_sid") or request.args.get("flowSid") or "").strip()

    q = Result.query
    if campaign_id:
        q = q.filter(Result.campaign_id == campaign_id)
    if flow_sid:
        q = q.filter(Result.flow_sid == flow_sid)

    rows = q.all()
    summary = {}
    total = 0
    for r in rows:
        total += 1
        try:
            answers = json.loads(r.answers_json or "{}")
        except Exception:
            answers = {}
        if not isinstance(answers, dict):
            continue
        for k, v in answers.items():
            k = str(k).lower()
            v = str(v)
            summary.setdefault(k, {})
            summary[k][v] = summary[k].get(v, 0) + 1

    return {"ok": True, "total_results": total, "summary": summary}


@app.get("/api/launches")
def api_launches():
    fail = require_passcode(request.args.to_dict(flat=True))
    if fail:
        return fail

    campaign_id = (request.args.get("campaign_id") or request.args.get("campaignId") or "").strip()
    flow_sid = (request.args.get("flow_sid") or request.args.get("flowSid") or "").strip()
    limit = int(request.args.get("limit", "200"))

    q = Launch.query.order_by(Launch.created_at.desc())
    if campaign_id:
        q = q.filter(Launch.campaign_id == campaign_id)
    if flow_sid:
        q = q.filter(Launch.flow_sid == flow_sid)

    rows = q.limit(min(limit, 2000)).all()

    execution_sids = [r.execution_sid for r in rows if r.execution_sid]
    outcome_map = {}
    answers_map_launches = {}
    call_sid_map_launches = {}

    if execution_sids:
        events = CallEvent.query.filter(CallEvent.execution_sid.in_(execution_sids)).all()
        for e in events:
            if e.execution_sid:
                outcome_map[e.execution_sid] = e.outcome
                if e.call_sid:
                    call_sid_map_launches[e.execution_sid] = e.call_sid
        res_rows = Result.query.filter(Result.execution_sid.in_(execution_sids)).all()
        for res in res_rows:
            try:
                ans = json.loads(res.answers_json or "{}")
            except Exception:
                ans = {}
            if isinstance(ans, dict):
                answers_map_launches[res.execution_sid] = ans

    duration_map_launches = {}
    if twilio_client and call_sid_map_launches:
        for exec_sid, csid in call_sid_map_launches.items():
            try:
                call = twilio_client.calls(csid).fetch()
                duration_map_launches[exec_sid] = int(call.duration or 0)
            except Exception:
                duration_map_launches[exec_sid] = None

    out = [{
        "id": r.id,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "campaign_id": r.campaign_id,
        "flow_sid": r.flow_sid,
        "execution_sid": r.execution_sid,
        "to": r.to_number,
        "from": r.from_number,
        "outcome": clasificar_estado(
            outcome_map.get(r.execution_sid),
            duration_map_launches.get(r.execution_sid),
            answers_map_launches.get(r.execution_sid, {})
        ),
    } for r in rows]

    return {"ok": True, "count": len(out), "launches": out}


@app.get("/api/twilio-calls")
def api_twilio_calls():
    fail = require_passcode(request.args.to_dict(flat=True))
    if fail:
        return fail

    if not twilio_client:
        return {"ok": False, "error": "Twilio no configurado."}, 500

    limit = int(request.args.get("limit", "50"))
    direction_filter = (request.args.get("direction", "outbound-api") or "").strip()
    include_inbound = (request.args.get("include_inbound", "false") or "").lower() == "true"

    def pick_attr(obj, *names):
        for n in names:
            v = getattr(obj, n, None)
            if v:
                return v
        return None

    try:
        kwargs = {"limit": min(limit, 200)}
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        if date_from:
            kwargs["start_time_after"] = datetime.strptime(date_from, "%Y-%m-%d")
        if date_to:
            kwargs["start_time_before"] = datetime.strptime(date_to, "%Y-%m-%d")
        calls = twilio_client.calls.list(**kwargs)
        items = []

        for c in calls:
            direction = pick_attr(c, "direction") or ""
            if not include_inbound and direction_filter:
                if direction != direction_filter:
                    continue

            from_val = (
                pick_attr(c, "from_", "from", "from_formatted", "caller", "caller_formatted")
                or (DEFAULT_FROM_NUMBER if direction == "outbound-api" else None)
            )
            to_val = pick_attr(c, "to", "to_formatted")
            start_time = pick_attr(c, "start_time")
            end_time = pick_attr(c, "end_time")

            items.append({
                "sid": c.sid,
                "from": from_val,
                "to": to_val,
                "status": pick_attr(c, "status"),
                "start_time": start_time.isoformat() if start_time else None,
                "end_time": end_time.isoformat() if end_time else None,
                "duration": pick_attr(c, "duration"),
                "direction": direction,
            })

        call_sids = [x.get("sid") for x in items if x.get("sid")]
        to_numbers = list({x.get("to") for x in items if x.get("to")})

        overrides = {}
        campaign_map = {}
        if call_sids:
            ce_rows = CallEvent.query.filter(CallEvent.call_sid.in_(call_sids)).all()
            overrides = {r.call_sid: r.outcome for r in ce_rows if r.call_sid and r.outcome}
            campaign_map = {r.call_sid: r.campaign_id for r in ce_rows if r.call_sid and r.campaign_id}

        result_campaign_map = {}
        if call_sids:
            res_rows = Result.query.filter(Result.call_sid.in_(call_sids)).all()
            result_campaign_map = {r.call_sid: r.campaign_id for r in res_rows if r.call_sid and r.campaign_id}

        # Cruzar answers por call_sid
        answers_by_call_sid = {}
        if call_sids:
            res_rows_answers = Result.query.filter(Result.call_sid.in_(call_sids)).all()
            for r in res_rows_answers:
                if r.call_sid:
                    try:
                        answers_by_call_sid[r.call_sid] = json.loads(r.answers_json or "{}")
                    except Exception:
                        answers_by_call_sid[r.call_sid] = {}

        launch_campaign_by_to = {}
        if to_numbers:
            launch_rows = (
                Launch.query
                .filter(Launch.to_number.in_(to_numbers))
                .order_by(Launch.created_at.desc())
                .all()
            )
            for lr in launch_rows:
                if lr.to_number and lr.to_number not in launch_campaign_by_to:
                    launch_campaign_by_to[lr.to_number] = lr.campaign_id

        # Obtener duraciones desde Twilio para clasificación inteligente
        duration_by_sid = {}
        if twilio_client and call_sids:
            for csid in call_sids:
                try:
                    call = twilio_client.calls(csid).fetch()
                    duration_by_sid[csid] = int(call.duration or 0)
                except Exception:
                    duration_by_sid[csid] = None

        for x in items:
            sid = x.get("sid")
            raw_outcome = overrides.get(sid) if sid else None
            answers = answers_by_call_sid.get(sid, {})
            duration = duration_by_sid.get(sid)

            # Aplicar clasificación inteligente
            x["status"] = clasificar_estado(
                raw_outcome or x.get("status"),
                duration,
                answers
            )
            campaign_id = (
                (campaign_map.get(sid, "") if sid else "")
                or (result_campaign_map.get(sid, "") if sid else "")
                or launch_campaign_by_to.get(x.get("to"), "")
            )
            x["campaign_id"] = campaign_id
            x["answers"] = answers
            x["duration"] = duration

        return {"ok": True, "count": len(items), "calls": items}

    except Exception as e:
        return {"ok": False, "error": str(e)}, 500


# =========================
# Routes — Exports
# =========================
@app.get("/download-results.xlsx")
def download_results_xlsx():
    fail = require_passcode(request.args.to_dict(flat=True))
    if fail:
        return fail

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import timedelta

    campaign_id = (request.args.get("campaign_id") or "").strip()
    q = Result.query.order_by(Result.created_at.desc())
    if campaign_id:
        q = q.filter(Result.campaign_id == campaign_id)

    rows = q.limit(5000).all()
    all_keys = set()
    parsed = []
    for r in rows:
        try:
            answers = json.loads(r.answers_json or "{}")
        except Exception:
            answers = {}
        if not isinstance(answers, dict):
            answers = {}
        all_keys.update(answers.keys())
        parsed.append((r, answers))

    qkeys = sorted([str(k) for k in all_keys])
    pregunta_headers = [f"Pregunta {k[1:]}" if str(k).startswith("q") and str(k)[1:].isdigit() else str(k) for k in qkeys]
    headers = ["Fecha", "Campaign ID", "Teléfono", "Flow SID", "Execution SID"] + pregunta_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (r, answers) in enumerate(parsed, 2):
        fecha = ""
        if r.created_at:
            mx_time = r.created_at - timedelta(hours=6)
            fecha = mx_time.strftime("%d/%m/%Y %H:%M:%S")
        row_data = [
            fecha, r.campaign_id or "", r.to_number or "",
            r.flow_sid or "", r.execution_sid or "",
        ] + [answers.get(k, "") for k in qkeys]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.font = Font(name="Arial", size=10)
            if col_idx == 3:
                cell.number_format = "@"

    col_widths = [22, 18, 18, 36, 36] + [14] * len(qkeys)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    fname = "results_{}.xlsx".format(campaign_id or "all")
    return send_file(mem, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/download-retry-list.xlsx")
def download_retry_list():
    """Exporta números no exitosos listos para relanzar."""
    fail = require_passcode(request.args.to_dict(flat=True))
    if fail:
        return fail

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import timedelta

    campaign_id = (request.args.get("campaign_id") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    q = Launch.query.order_by(Launch.created_at.desc())
    if campaign_id:
        q = q.filter(Launch.campaign_id == campaign_id)
    if date_from:
        try:
            q = q.filter(Launch.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        except Exception:
            pass
    if date_to:
        try:
            q = q.filter(Launch.created_at <= datetime.strptime(date_to, "%Y-%m-%d"))
        except Exception:
            pass

    rows = q.limit(10000).all()
    execution_sids = [r.execution_sid for r in rows if r.execution_sid]

    outcome_map = {}
    answers_map = {}
    call_sid_map = {}

    if execution_sids:
        events = CallEvent.query.filter(CallEvent.execution_sid.in_(execution_sids)).all()
        for e in events:
            if e.execution_sid:
                outcome_map[e.execution_sid] = e.outcome
                if e.call_sid:
                    call_sid_map[e.execution_sid] = e.call_sid
        res_rows = Result.query.filter(Result.execution_sid.in_(execution_sids)).all()
        for res in res_rows:
            try:
                ans = json.loads(res.answers_json or "{}")
            except Exception:
                ans = {}
            if isinstance(ans, dict):
                answers_map[res.execution_sid] = ans

    duration_map = {}
    if twilio_client and call_sid_map:
        for exec_sid, csid in call_sid_map.items():
            try:
                call = twilio_client.calls(csid).fetch()
                duration_map[exec_sid] = int(call.duration or 0)
            except Exception:
                duration_map[exec_sid] = None

    retry_rows = []
    for r in rows:
        estado = clasificar_estado(
            outcome_map.get(r.execution_sid),
            duration_map.get(r.execution_sid),
            answers_map.get(r.execution_sid, {})
        )
        if estado in ESTADOS_RELANZAR:
            retry_rows.append((r, estado))

    wb = Workbook()
    ws = wb.active
    ws.title = "Numeros a Relanzar"

    ws.merge_cells("A1:D1")
    nota = ws.cell(row=1, column=1,
        value="INSTRUCCIONES: Copia la columna A (Telefono) y pegala en Outbound para relanzar. Agrega -RETRY-1 al Campaign ID.")
    nota.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    nota.fill = PatternFill("solid", start_color="1F7A4A")
    nota.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 30

    headers = ["Telefono", "Motivo no exitosa", "Fecha del intento", "Campaign ID"]
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (r, estado) in enumerate(retry_rows, 3):
        fecha = ""
        if r.created_at:
            mx_time = r.created_at - timedelta(hours=6)
            fecha = mx_time.strftime("%d/%m/%Y %H:%M")
        c1 = ws.cell(row=row_idx, column=1, value=str(r.to_number or ""))
        c1.font = Font(name="Arial", size=10)
        c1.number_format = "@"
        ws.cell(row=row_idx, column=2, value=estado).font = Font(name="Arial", size=10)
        ws.cell(row=row_idx, column=3, value=fecha).font = Font(name="Arial", size=10)
        ws.cell(row=row_idx, column=4, value=r.campaign_id or "").font = Font(name="Arial", size=10)

    for i, w in enumerate([20, 28, 22, 25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    fname = "relanzar_{}.xlsx".format(campaign_id or "all")
    return send_file(mem, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def translate_status(outcome, answers):
    if not outcome or outcome in ("no-answer", "no_answer"):
        return "No contestó"
    if outcome == "busy":
        return "Ocupado"
    if outcome == "failed":
        return "Fallida"
    if outcome == "completed":
        if not answers:
            return "No contestó"
        values = list(answers.values())
        all_empty = all(not v or v.strip() == "" or v == "Sin respuesta" for v in values)
        return "No contestó" if all_empty else "Completada"
    return "No contestó"


@app.get("/download-launches.csv")
def download_launches_csv():
    fail = require_passcode(request.args.to_dict(flat=True))
    if fail:
        return fail

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import timedelta
        use_xlsx = True
    except ImportError:
        use_xlsx = False

    campaign_id = (request.args.get("campaign_id") or "").strip()
    q = Launch.query.order_by(Launch.created_at.desc())
    if campaign_id:
        q = q.filter(Launch.campaign_id == campaign_id)

    rows = q.limit(10000).all()
    execution_sids = [r.execution_sid for r in rows if r.execution_sid]

    outcome_map = {}
    if execution_sids:
        events = CallEvent.query.filter(CallEvent.execution_sid.in_(execution_sids)).all()
        for e in events:
            if e.execution_sid:
                outcome_map[e.execution_sid] = e.outcome

    answers_map = {}
    all_qkeys = set()
    if execution_sids:
        results = Result.query.filter(Result.execution_sid.in_(execution_sids)).all()
        for res in results:
            try:
                ans = json.loads(res.answers_json or "{}")
            except Exception:
                ans = {}
            if isinstance(ans, dict):
                all_qkeys.update(ans.keys())
                answers_map[res.execution_sid] = ans

    qkeys = sorted([str(k) for k in all_qkeys])

    call_sid_map = {}
    if execution_sids:
        ce_rows = CallEvent.query.filter(CallEvent.execution_sid.in_(execution_sids)).all()
        call_sid_map = {e.execution_sid: e.call_sid for e in ce_rows if e.call_sid}

    duration_map = {}
    if twilio_client and call_sid_map:
        for exec_sid, csid in call_sid_map.items():
            try:
                call = twilio_client.calls(csid).fetch()
                duration_map[exec_sid] = int(call.duration or 0)
            except Exception:
                duration_map[exec_sid] = ""

    headers = ["Fecha", "Campaign ID", "De", "Para", "Estatus", "Duracion (seg)", "Execution SID"] + qkeys

    if use_xlsx:
        from datetime import timedelta
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        header_fill = PatternFill("solid", start_color="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, r in enumerate(rows, 2):
            fecha = ""
            if r.created_at:
                mx_time = r.created_at - timedelta(hours=6)
                fecha = mx_time.strftime("%d/%m/%Y %H:%M:%S")
            estatus = clasificar_estado(outcome_map.get(r.execution_sid), duration_map.get(r.execution_sid), answers_map.get(r.execution_sid, {}))
            duracion = duration_map.get(r.execution_sid, "")
            ans = answers_map.get(r.execution_sid, {})
            row_data = [
                fecha, r.campaign_id, r.from_number or "", r.to_number or "",
                estatus, duracion, r.execution_sid or "",
            ] + [ans.get(k, "") for k in qkeys]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
                cell.font = Font(name="Arial", size=10)
                if col_idx in (3, 4):
                    cell.number_format = "@"

        col_widths = [22, 18, 18, 18, 14, 14, 36] + [14] * len(qkeys)
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        fname = "historial_{}.xlsx".format(campaign_id or "all")
        return send_file(mem, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    from datetime import timedelta
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for r in rows:
        fecha = ""
        if r.created_at:
            mx_time = r.created_at - timedelta(hours=6)
            fecha = mx_time.strftime("%d/%m/%Y %H:%M:%S")
        ans = answers_map.get(r.execution_sid, {})
        estatus = clasificar_estado(outcome_map.get(r.execution_sid), duration_map.get(r.execution_sid), ans)
        duracion = duration_map.get(r.execution_sid, "")
        writer.writerow([
            fecha, r.campaign_id, r.from_number or "", r.to_number or "",
            estatus, duracion, r.execution_sid or "",
        ] + [ans.get(k, "") for k in qkeys])
    mem = io.BytesIO(output.getvalue().encode("utf-8"))
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name="historial.csv", mimetype="text/csv")


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=True)
